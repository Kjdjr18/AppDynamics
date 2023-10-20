import requests
import xml.etree.ElementTree as ET
import certifi
import os
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
os.environ['REQUESTS_CA_BUNDLE'] = certifi.where()

# Determine the root folder of the script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Create a 'output' directory if it doesn't exist
output_dir = os.path.join(script_dir, 'output')
os.makedirs(output_dir, exist_ok=True)

# Define the path for the output Excel file
filename = os.path.join(output_dir, 'Applications.xlsx')


# Replace with your AppDynamics controller info and API token
ACCOUNT_NAME = "ACCOUNT HERE"  # account name here
DEFAULT_HOST = f"https://{ACCOUNT_NAME}.saas.appdynamics.com"
CONTROLLER_HOST = DEFAULT_HOST  # replace with host (default saas)
CONTROLLER_PORT = 443  # add port for host (443 for saas)
API_TOKEN = "API KEY HERE"  # add api key here

APPLICATION_LIST = ["app1, app2"] # add apps list here

# LEAVE THE APPLICATION ID BLANK
APPLICATION_ID = ""


# Get them app names
def get_application_name(application_id):
    url = f"https://{CONTROLLER_HOST}:{CONTROLLER_PORT}/controller/rest/applications/{application_id}"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_TOKEN}"
    }
    try:
        response = requests.get(url, headers=headers, verify=False)
        response.raise_for_status()  # Raises a HTTPError if the response has an error status code

        # Check if response body is not empty
        if response.text.strip():
            # Parse the XML response
            root = ET.fromstring(response.content)

            # Search for the 'application' tags in the XML, then find the 'name' child
            # and compare the 'id' child text with the desired application_id.
            for application in root.findall('application'):
                id_ = application.find('id').text
                if id_ == str(application_id):
                    name = application.find('name').text
                    return name

            print(f"No application with ID: {application_id}")
            return None

        else:
            print(f"Empty response received for application ID: {application_id}")
            return None

    except requests.exceptions.HTTPError as e:
        print(f"HTTP error occurred for application ID {application_id}: {e}")
        print(response.text)  # Print the actual response content, which might include the error message
        return None

    except ET.ParseError:
        print("Error parsing XML. Response was not valid XML.")
        print(response.text)  # Print the actual response content
        return None

    except requests.exceptions.RequestException as e:
        print(f"An error occurred: {e}")
        return None


# Get them Tiers
def get_tiers(application_id):
    url = f"https://{CONTROLLER_HOST}:{CONTROLLER_PORT}/controller/rest/applications/{application_id}/tiers"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_TOKEN}"
    }
    params = {
        "output": "JSON"
    }
    try:
        response = requests.get(url, headers=headers, params=params, verify=False)
        if response.status_code == 200:
            data = response.json()
            tier_names = [tier['name'] for tier in data]
            return tier_names
        else:
            print(f"Failed to fetch tiers. Status code: {response.status_code}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error connecting to the AppDynamics API: {e}")
        return None


# Get them backends
def get_backends(application_id):
    url = f"https://{CONTROLLER_HOST}:{CONTROLLER_PORT}/controller/rest/applications/{application_id}/backends"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_TOKEN}"
    }
    params = {
        "output": "JSON"
    }
    try:
        response = requests.get(url, headers=headers, params=params, verify=False)
        if response.status_code == 200:
            data = response.json()
            backend_names = [backend['name'] for backend in data]
            return backend_names
        else:
            print(f"Failed to fetch backends for tier '{application_id}'. Status code: {response.status_code}")
            return None
    except requests.exceptions.RequestException as e:
        print(f"Error connecting to the AppDynamics API: {e}")
        return None


# Function that structures and outputs stuff to the console
def structure_application_data(application_id, app_name, tiers, backends):
    """
    Structures the application data into a dictionary.

    :param application_id: ID of the application.
    :param app_name: Name of the application.
    :param tiers: List of tiers.
    :param backends: List of backends.
    :return: Dictionary containing structured application data.
    """
    return {
        'Application ID': application_id,
        'Application Name': app_name,
        'Tiers': tiers or ['N/A'],  # Defaulting to ['N/A'] if tiers list is empty
        'Backends': backends or ['N/A']  # Defaulting to ['N/A'] if backends list is empty
    }


def collect_and_structure_app_data(application_list):
    structured_data = {}

    for application_id in application_list:
        app_name = get_application_name(application_id)
        if not app_name:
            print(f"Failed to retrieve data for Application ID: {application_id}")
            continue

        tiers = get_tiers(application_id)
        backends = get_backends(application_id)

        app_data = structure_application_data(application_id, app_name, tiers, backends)
        app_key = f"{app_name} ({application_id})"
        structured_data[app_key] = app_data

    return structured_data


def print_application_data(structured_data):
    for app_key, app_details in structured_data.items():
        print(f"\nApplication Name: {app_details['Application Name']} (ID: {app_details['Application ID']})\n")

        print("Tiers:")
        for tier in app_details['Tiers']:
            print(f"  {tier}")

        print("\nBackends:")
        for backend in app_details['Backends']:
            print(f"  {backend}")

        print('\n' + '-' * 30 + '\n')  # Separator for readability


def create_excel_file(structured_data, filename):
    workbook = Workbook()
    sheet_summary = workbook.active
    sheet_summary.title = 'Applications Summary'

    # Formatting
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    aligned_center = Alignment(horizontal='center', vertical='center')

    # Headers for "Application Name (ID)," "Tiers," and "Backends"
    sheet_summary['A1'] = 'Application Name (ID)'
    sheet_summary['B1'] = 'Tiers'
    sheet_summary['C1'] = 'Backends'

    # Apply borders and center alignment
    for cell in ['A1', 'B1', 'C1']:
        sheet_summary[cell].alignment = aligned_center
        sheet_summary[cell].border = thin_border

    row_num = 2  # Starting row for data
    for app_key, app_details in structured_data.items():
        tiers = '\n'.join(app_details['Tiers'])  # Convert list to string with new lines
        backends = '\n'.join(app_details['Backends'])  # Convert list to string with new lines

        # Write data in the correct columns
        sheet_summary.cell(row=row_num, column=1, value=app_key).border = thin_border
        sheet_summary.cell(row=row_num, column=2, value=tiers).border = thin_border
        sheet_summary.cell(row=row_num, column=3, value=backends).border = thin_border

        # Adjusting the row height for better view
        sheet_summary.row_dimensions[row_num].height = max(15, len(tiers.split('\n')) * 15,
                                                           len(backends.split('\n')) * 15)
        row_num += 1

    # Auto-sizing columns for better view
    for column in sheet_summary.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet_summary.column_dimensions[column[0].column_letter].width = adjusted_width

    # Save the workbook
    workbook.save(filename)


if __name__ == "__main__":
    application_list = APPLICATION_LIST  # Replace with your actual list of application IDs

    print("Processing application data...\n")
    structured_data = collect_and_structure_app_data(application_list)

    print("Printing application data...\n")
    print_application_data(structured_data)

    if structured_data:
        print("\nWriting data to Excel file...")
        create_excel_file(structured_data, 'output/Applications.xlsx')
        print("Data has been written to Excel successfully.")
    else:
        print("No data available to write to Excel.")